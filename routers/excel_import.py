"""
Excel Import Router
Handles bulk employee imports from Excel files
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from datetime import datetime, date
from database import get_db
from auth import get_current_user
import models
import openpyxl
from io import BytesIO
import logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/excel", tags=["excel"])


@router.post("/import-employees")
async def import_employees_from_excel(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Bulk import employees from Excel file.
    Expected columns:
    - file_code (required)
    - full_name (required)
    - email (required)
    - contact_number
    - position (required)
    - project
    - location
    - employment_type
    - date_of_appointment
    - contract_start
    - contract_end
    """
    if current_user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admins can import employees")
    
    try:
        # Read Excel file
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        worksheet = workbook.active
        
        # Parse headers
        headers = {}
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value:
                headers[cell.value.lower().strip()] = col_idx
        
        required_fields = ["file_code", "full_name", "email", "position"]
        for field in required_fields:
            if field not in headers:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required column: {field}"
                )
        
        imported_employees = []
        errors = []
        
        # Import rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), 2):
            try:
                # Extract values
                file_code = row[headers["file_code"] - 1].value
                full_name = row[headers["full_name"] - 1].value
                email = row[headers["email"] - 1].value
                position = row[headers["position"] - 1].value
                
                if not all([file_code, full_name, email, position]):
                    errors.append(f"Row {row_idx}: Missing required fields")
                    continue
                
                # Get optional fields
                contact_number = row[headers.get("contact_number", 0) - 1].value if "contact_number" in headers else None
                project = row[headers.get("project", 0) - 1].value if "project" in headers else None
                location = row[headers.get("location", 0) - 1].value if "location" in headers else None
                employment_type = row[headers.get("employment_type", 0) - 1].value if "employment_type" in headers else "Contract"
                
                # Parse dates
                date_of_appointment = None
                contract_start = None
                contract_end = None
                
                if "date_of_appointment" in headers:
                    doa = row[headers["date_of_appointment"] - 1].value
                    if doa:
                        date_of_appointment = doa if isinstance(doa, date) else datetime.strptime(str(doa), "%Y-%m-%d").date()
                
                if "contract_start" in headers:
                    cs = row[headers["contract_start"] - 1].value
                    if cs:
                        contract_start = cs if isinstance(cs, date) else datetime.strptime(str(cs), "%Y-%m-%d").date()
                
                if "contract_end" in headers:
                    ce = row[headers["contract_end"] - 1].value
                    if ce:
                        contract_end = ce if isinstance(ce, date) else datetime.strptime(str(ce), "%Y-%m-%d").date()
                
                # Check if employee already exists
                existing = db.query(models.Employee).filter(
                    models.Employee.file_code == str(file_code)
                ).first()
                
                if existing:
                    # Update existing employee
                    existing.full_name = full_name
                    existing.position = position
                    existing.contact_number = contact_number
                    existing.project = project
                    existing.location = location
                    existing.employment_type = employment_type
                    if date_of_appointment:
                        existing.date_of_appointment = date_of_appointment
                    if contract_start:
                        existing.contract_start = contract_start
                    if contract_end:
                        existing.contract_end = contract_end
                    existing.status = "Active"
                    db.add(existing)
                    imported_employees.append({
                        "file_code": file_code,
                        "name": full_name,
                        "status": "updated"
                    })
                else:
                    # Create new employee
                    employee = models.Employee(
                        file_code=str(file_code),
                        full_name=full_name,
                        position=position,
                        contact_number=contact_number,
                        project=project,
                        location=location,
                        employment_type=employment_type,
                        date_of_appointment=date_of_appointment,
                        contract_start=contract_start,
                        contract_end=contract_end,
                        status="Active"
                    )
                    db.add(employee)
                    imported_employees.append({
                        "file_code": file_code,
                        "name": full_name,
                        "status": "created"
                    })
            
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                logger.error(f"Error importing row {row_idx}: {str(e)}")
        
        # Commit all changes
        db.commit()
        
        return {
            "success": True,
            "message": f"Imported {len(imported_employees)} employees",
            "imported": imported_employees,
            "errors": errors,
            "total_rows_processed": len(list(worksheet.iter_rows(min_row=2))) if worksheet.max_row > 1 else 0
        }
    
    except openpyxl.utils.exceptions.InvalidFileException:
        raise HTTPException(status_code=400, detail="Invalid Excel file format")
    except Exception as e:
        logger.error(f"Error importing Excel file: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error importing file: {str(e)}")


@router.get("/employee-template")
async def get_employee_import_template(current_user: models.User = Depends(get_current_user)):
    """
    Get Excel template for employee import.
    Returns a sample Excel file with required and optional columns.
    """
    if current_user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admins can access templates")
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"
    
    # Add headers
    headers = [
        "file_code",
        "full_name",
        "email",
        "position",
        "contact_number",
        "project",
        "location",
        "employment_type",
        "date_of_appointment",
        "contract_start",
        "contract_end"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E5A96", end_color="1E5A96", fill_type="solid")
    
    # Add sample row
    sample_data = [
        "EMP001",
        "John Doe",
        "john@example.com",
        "Senior Manager",
        "+256-123-456789",
        "Project Alpha",
        "Kampala",
        "Permanent",
        "2023-01-15",
        "2023-01-15",
        "2026-01-14"
    ]
    
    for col_idx, value in enumerate(sample_data, 1):
        worksheet.cell(row=2, column=col_idx).value = value
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return {
        "message": "Template created. Use columns: file_code (required), full_name (required), email (required), position (required), and optional columns for other data.",
        "required_columns": ["file_code", "full_name", "email", "position"],
        "optional_columns": ["contact_number", "project", "location", "employment_type", "date_of_appointment", "contract_start", "contract_end"]
    }
